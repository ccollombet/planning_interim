import streamlit as st
import pandas as pd
import os
import tempfile
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy
import re
from datetime import datetime

# ✅ CECI DOIT ÊTRE LA PREMIÈRE COMMANDE STREAMLIT
st.set_page_config(page_title="Générateur de planning", layout="centered")
# === CONFIGURATION ===

PASSWORD = st.secrets["PLANNING_APP_PASSWORD"]



# === AUTHENTIFICATION ===
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        with st.form("Login"):
            pwd = st.text_input("🔐 Entrez le mot de passe", type="password")
            submitted = st.form_submit_button("Se connecter")
            if submitted and pwd == PASSWORD:
                st.session_state.authenticated = True
            elif submitted:
                st.error("Mot de passe incorrect")
    return st.session_state.authenticated

if not check_password():
    st.stop()

#st.set_page_config(page_title="Générateur de planning", layout="centered")
st.title("🗓️ Générateur de planning MAS Montaines")

# === Fonctions utilitaires génériques ===
def save_uploaded_file(uploaded_file, suffix):
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, f"{Path(uploaded_file.name).stem}_{suffix}{Path(uploaded_file.name).suffix}")
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return file_path

def traitement_partie1(fichier_initial):
    fichier_csv = "fichier_intermediaire.csv"
    fichier_nettoye = "planning_filtre.xlsx"
    fichier_nom_prenom = "planning_avec_nom_prenom.xlsx"
    fichier_final = "planning_final_complet.xlsx"

    wb = load_workbook(fichier_initial)
    ws = wb.active
    row_data, dernier_groupe = [], None
    dans_bloc_rempla = False

    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        val = val.strip() if isinstance(val, str) else ""

        if val.lower().startswith("rempla"):
            dernier_groupe = val.replace("\n", " ").strip()
            dans_bloc_rempla = True
        elif re.match(r"^[A-ZÉÈÀÙÂÊÎÔÛÇ]+(\s+[A-ZÉÈÀÙÂÊÎÔÛÇ]+)*$", val):
            dans_bloc_rempla = False
            dernier_groupe = None
        elif re.match(r"\d{2}/\d{2}/\d{4}", val) and dans_bloc_rempla and dernier_groupe:
            try:
                date_str, nom_prenom = val.split(":", 1)
                nom_parts = nom_prenom.strip().split()
                if len(nom_parts) >= 2:
                    nom = nom_parts[0]
                    prenom = " ".join(nom_parts[1:])
                    mots_exclus = ["planning", "modif", "absence", "réunion", "rdv", "formation",
                                   "pass", "prévu", "anticipé", "mains", "demande", "chgt",
                                   "en", "lieu", "possible", "heures", "déplacement", "dt", "cadre", "ok"]
                    if not any(mot in prenom.lower() for mot in mots_exclus):
                        row_data.append({"date": date_str.strip(), "groupe": dernier_groupe, "nom": nom, "prenom": prenom})
            except Exception:
                pass
    pd.DataFrame(row_data).to_csv(fichier_csv, index=False)

    wb = load_workbook(fichier_initial)
    ws = wb.active
    wb_nouveau = Workbook()
    ws_nouveau = wb_nouveau.active
    ligne_nouvelle = 1

    for row in ws.iter_rows():
        val = row[0].value
        if isinstance(val, str) and re.match(r"\d{2}/\d{2}/\d{4}", val.strip()):
            continue
        for col_index, cell in enumerate(row, start=1):
            nc = ws_nouveau.cell(row=ligne_nouvelle, column=col_index, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.border = copy(cell.border)
                nc.fill = copy(cell.fill)
                nc.number_format = copy(cell.number_format)
                nc.protection = copy(cell.protection)
                nc.alignment = copy(cell.alignment)
        ligne_nouvelle += 1

    wb_nouveau.save(fichier_nettoye)
    wb = load_workbook(fichier_nettoye)
    ws = wb.active
    lignes_act_jour = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 3).value == "Act. jour"]
    decalage = 0
    for ligne in lignes_act_jour:
        i = ligne + 1 + decalage
        ws.insert_rows(i, amount=2)
        ws.cell(row=i, column=3, value="Nom").font = Font(name="Segoe UI", size=14)
        ws.cell(row=i+1, column=3, value="Prénom").font = Font(name="Segoe UI", size=14)
        decalage += 2

    wb.save(fichier_nom_prenom)
    df_rempla = pd.read_csv(fichier_csv)
    wb = load_workbook(fichier_nom_prenom)
    ws = wb.active
    colonnes = range(4, 35)
    mois_map = {"Jan": "01", "Fév": "02", "Fev": "02", "Mar": "03", "Avr": "04", "Mai": "05", "Juin": "06",
                "Juil": "07", "Aoû": "08", "Aou": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Déc": "12", "Dec": "12"}

    def convertir(cellule):
        if isinstance(cellule, str) and "\n" in cellule:
            j, m = cellule.strip().split("\n")
            j = j[1:] if j[0] in "LMJVSD" else j
            #return f"{int(j):02d}/{mois_map.get(m[:3], '00')}/2025"
            return f"{int(j):02d}/{mois_map.get(m.strip().capitalize()[:4], '00')}/2025"
        return None

    #for row in range(1, ws.max_row - 4):
    for row in range(1, ws.max_row):
        if ws.cell(row=row, column=1).value:  # force le traitement si la colonne A est remplie

        #if ws.cell(row=row, column=3).value == "Hor.":
            r_hor, r_nom, r_pre = row, row+3, row+4
            identite = ws.cell(r_hor, column=1).value
            if not isinstance(identite, str): continue
            is_rempla = identite.strip().lower().startswith("rempla")
            groupe = identite.replace("\n", " ").strip()

            ##if not is_rempla and "\n" in identite:
            #if not is_rempla and isinstance(identite, str) and len(identite.splitlines()) == 2:
#
            #    #nom, prenom = identite.split("\n", 1)
            #    nom, prenom = identite.splitlines()
#
            #    ws.cell(r_hor, 1, f"{nom.strip()}\n{prenom.strip()}").alignment = Alignment(wrap_text=True)
            #    for col in colonnes:
            #        for r, val in zip([r_nom, r_pre], [nom.strip(), prenom.strip()]):
            #            c = ws.cell(r, col, val)
            #            c.font = Font(name="Segoe UI", size=8)
            #            c.fill = copy(ws.cell(r_hor + 2, col).fill)
            #            c.alignment = Alignment(horizontal="center")
            if not is_rempla:
                nom_prenom_parts = identite.strip().split()
                if len(nom_prenom_parts) >= 2:
                    nom = nom_prenom_parts[0].strip()
                    prenom = " ".join(nom_prenom_parts[1:]).strip()
                    ws.cell(r_hor, 1, f"{nom}\n{prenom}").alignment = Alignment(wrap_text=True)
                    for col in colonnes:
                        for r, val in zip([r_nom, r_pre], [nom, prenom]):
                            c = ws.cell(r, col, val)
                            c.font = Font(name="Segoe UI", size=8)
                            c.fill = copy(ws.cell(r_hor + 2, col).fill)
                            c.alignment = Alignment(horizontal="center")

            elif is_rempla:
                for col in colonnes:
                    d = convertir(ws.cell(1, col).value)
                    match = df_rempla[(df_rempla["groupe"].str.strip() == groupe) & (df_rempla["date"].str.strip() == d)]
                    if not match.empty:
                        nom_csv, prenom_csv = match.iloc[0]["nom"], match.iloc[0]["prenom"]
                        for r, val in zip([r_nom, r_pre], [nom_csv, prenom_csv]):
                            c = ws.cell(r, col, val)
                            c.font = Font(name="Segoe UI", size=8)
                            c.fill = copy(ws.cell(r_hor + 2, col).fill)
                            c.alignment = Alignment(horizontal="center")

    for row in range(1, ws.max_row + 1):
        val_C = ws.cell(row, 3).value
        if isinstance(val_C, str) and val_C.strip() == "Hor.":
            max_ligne = 40
            for col in range(4, 35):
                cell = ws.cell(row, col)
                if isinstance(cell.value, str):
                    txt = re.sub(r"\s*-\s*", " -\n", cell.value.strip())
                    txt = txt.replace("/", "/\n")
                    cell.value = txt
                    cell.alignment = Alignment(wrap_text=True, horizontal="center")
                    if "/\n" in txt or txt.count("\n") > 1:
                        max_ligne = 80
            ws.row_dimensions[row].height = max_ligne

    wb_new = Workbook()
    ws_new = wb_new.active
    r_new = 1
    for row in ws.iter_rows(min_col=1, max_col=34):
        if all((cell.value in [None, ""]) for cell in row): continue
        for col_index, cell in enumerate(row, start=1):
            nc = ws_new.cell(row=r_new, column=col_index, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.border = copy(cell.border)
                nc.fill = copy(cell.fill)
                nc.number_format = copy(cell.number_format)
                nc.protection = copy(cell.protection)
                nc.alignment = copy(cell.alignment)
        r_new += 1

    for row in range(1, ws_new.max_row - 3):
        if ws_new.cell(row=row, column=3).value == "Hor.":
            ws_new.merge_cells(start_row=row, end_row=row + 4, start_column=1, end_column=1)
            ws_new.merge_cells(start_row=row, end_row=row + 4, start_column=2, end_column=2)
            for col in [1, 2]:
                cell = ws_new.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Ajuster la largeur de la colonne A
    ws_new.column_dimensions["A"].width = 50

    wb_new.save(fichier_final)
    return fichier_final

# === Partie 1 : Upload et génération planning stylisé ===
st.header("1️⃣ Upload et génération du planning stylisé")
uploaded_file_1 = st.file_uploader("Uploader le planning brut", type=["xlsx"], key="upload1")
if uploaded_file_1 and st.button("Générer le planning stylisé"):
    raw_path = save_uploaded_file(uploaded_file_1, "raw")
    with st.spinner("Traitement du fichier en cours..."):
        fichier_final = traitement_partie1(raw_path)
    st.success("✅ Planning stylisé généré !")
    st.download_button("📥 Télécharger le fichier final", data=open(fichier_final, "rb"), file_name=os.path.basename(fichier_final))

def load_dataframe_from_excel(fichier):
    return pd.read_excel(fichier, sheet_name=None)

def traitement_partie2(fichier_source):
    wb = load_workbook(fichier_source)
        # 1️⃣ : supprimer toutes les tables héritées du classeur source
    for ws in wb.worksheets:
        for tbl in list(ws._tables):
            ws.remove_table(tbl.name)

    # 2️⃣ : supprimer d'éventuels onglets "lecture" / "interimaire"
    for name in ["lecture", "interimaire"]:
        if name in wb.sheetnames:
            wb.remove(wb[name])
    #for ws in wb.worksheets:
    #    ws._tables = []
#
    #for name in ["lecture", "interimaire"]:
    #    if name in wb.sheetnames:
    #        std = wb[name]
    #        wb.remove(std)
#
    ws_source = wb.active
    colonnes = list(range(4, 35))
    mois_map = {
        "Jan": "01", "Fév": "02", "Fev": "02", "Mar": "03", "Mars": "03", "Avr": "04", "Mai": "05",
        "Juin": "06", "Jui": "07", "Juil": "07", "Août": "08", "Aou": "08", "Sep": "09", "Sept": "09",
        "Oct": "10", "Nov": "11", "Déc": "12", "Dec": "12"
    }

    dates_colonnes = {}
    for col in colonnes:
        val = ws_source.cell(row=1, column=col).value
        if isinstance(val, str) and "\n" in val:
            parts = val.strip().split("\n")
            if len(parts) == 2:
                jour, mois = parts
                jour = jour[1:] if jour and jour[0] in "LMJVSD" else jour
                if jour.isdigit():
                    dates_colonnes[col] = f"{int(jour):02d}/{mois_map.get(mois.strip().capitalize()[:4], '00')}/2025"

    lignes_donnees = []
    #for row in range(1, ws_source.max_row - 4):
    #    if ws_source.cell(row=row, column=3).value == "Hor.":
    #        ligne_hor = row
    #        ligne_lieu = row + 1
    #        ligne_act = row + 2
    #        ligne_nom = row + 3
    #        ligne_prenom = row + 4
#
    #        valeur_nom_colA = ws_source.cell(row=ligne_hor, column=1).value or ""
#
    #        for col in colonnes:
    #            val_act = ws_source.cell(row=ligne_act, column=col).value
    #            val_hor = ws_source.cell(row=ligne_hor, column=col).value
    #            val_lieu = ws_source.cell(row=ligne_lieu, column=col).value
#
    #            if isinstance(val_act, str) and val_act.startswith("502G"):
    #                groupe = val_act[-1] if val_act[-1].isdigit() else ""
    #                date = dates_colonnes.get(col, "")
    #                lignes_donnees.append([
    #                    date, groupe, val_hor, "", "", valeur_nom_colA.replace("\n", " "), val_lieu
    #                ])

       # ➡️ Au lieu de faire `for row in range(1, ws_source.max_row - 4):`
    # on va détecter **toutes** les lignes où col=3 vaut "Hor."
    max_row = ws_source.max_row
    lignes_hor = [
        r for r in range(1, max_row + 1)
        if ws_source.cell(row=r, column=3).value == "Hor."
    ]

    lignes_donnees = []
    for ligne_hor in lignes_hor:
        # on s'assure qu'on a bien la place pour lire les 4 lignes suivantes
        if ligne_hor + 4 > max_row:
            # si on est trop près de la fin, on saute
            continue

        ligne_lieu   = ligne_hor + 1
        ligne_act    = ligne_hor + 2
        ligne_nom    = ligne_hor + 3
        ligne_prenom = ligne_hor + 4

        valeur_nom_colA = ws_source.cell(row=ligne_hor, column=1).value or ""

        for col in colonnes:
            val_act = ws_source.cell(row=ligne_act, column=col).value
            val_hor = ws_source.cell(row=ligne_hor, column=col).value
            val_lieu = ws_source.cell(row=ligne_lieu, column=col).value

            if isinstance(val_act, str) and val_act.startswith("502G"):
                groupe = val_act[-1] if val_act[-1].isdigit() else ""
                date   = dates_colonnes.get(col, "")
                lignes_donnees.append([
                    date,
                    groupe,
                    val_hor,
                    "",  # Motif
                    "",  # NOM de la personne remplacée
                    valeur_nom_colA.replace("\n", " "),
                    val_lieu
                ])

    lignes_donnees = sorted(lignes_donnees, key=lambda x: (pd.to_datetime(x[0], dayfirst=True), x[1]))

    entetes = ["Date", "Groupe", "Horaire", "Motif", "NOM de la personne remplacée", "Nom", "Agence"]

    def creer_onglet(nom_onglet, lignes):
        ws = wb.create_sheet(nom_onglet)
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=9)
        cell_titre = ws.cell(row=3, column=3, value="MAS MONTAINES")
        cell_titre.alignment = Alignment(horizontal="center")
        cell_titre.font = Font(name="Aptos Narrow", size=11, bold=True)
        cell_titre.fill = PatternFill(start_color="FBE2D5", end_color="FBE2D5", fill_type="solid")

        for idx, val in enumerate(entetes, start=3):
            cell = ws.cell(row=4, column=idx, value=val)
            cell.font = Font(name="Aptos Narrow", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center")

        ligne = 5
        for ligne_data in lignes:
            for col_index, val in enumerate(ligne_data, start=3):
                cell = ws.cell(row=ligne, column=col_index, value=val)
                cell.font = Font(name="Aptos Narrow", size=11)
                cell.alignment = Alignment(horizontal="center")
            ligne += 1

        if lignes:
            ref = f"C4:I{ligne - 1}"
            table = Table(displayName=f"Table{nom_onglet.capitalize()}", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

        for col_idx in range(3, 10):
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = 28 if nom_onglet == "lecture" else 20

    creer_onglet("lecture", lignes_donnees)

    df = pd.DataFrame(lignes_donnees, columns=entetes)
    df_interim = df[df["Agence"] == "A POURVOIR"].copy()
    df_interim["Date"] = pd.to_datetime(df_interim["Date"], dayfirst=True)
    df_interim = df_interim.sort_values(by=["Date", "Groupe"])
    df_interim["Date"] = df_interim["Date"].dt.strftime("%d/%m/%Y")
    creer_onglet("interimaire", df_interim.values.tolist())

    fichier_sortie = "planning_final_complet.xlsx"
    wb.save(fichier_sortie)
    return fichier_sortie

st.header("2️⃣ Génération des onglets 'lecture' et 'interimaire'")
uploaded_file_2 = st.file_uploader("Uploader le fichier planning stylisé modifié", type=["xlsx"], key="upload2")
if uploaded_file_2 and st.button("Créer les onglets Lecture & Interimaire"):
    stylized_path = save_uploaded_file(uploaded_file_2, "stylized")
    with st.spinner("Création des onglets Lecture et Interimaire..."):
        fichier_resultat = traitement_partie2(stylized_path)
    st.success("✅ Onglets créés !")
    st.download_button("📅 Télécharger le fichier modifié", data=open(fichier_resultat, "rb"), file_name=os.path.basename(fichier_resultat))


# === Partie 3 : Génération du fichier Badakan ===
def traitement_partie3(fichier_interimaire):
    df_interim = pd.read_excel(fichier_interimaire, sheet_name="interimaire", skiprows=3, usecols="C:I")
    df_interim.columns = ["Date", "Groupe", "Horaire", "Motif", "Personne remplacée", "Nom", "Agence"]


    horaire_mapping = {
        "700-1430": (7.5, 0), "1400-2130": (7.5, 0), "800-2000": (12, 0), "0700-1430": (7.5, 0),  "0800-2000": (12, 0),
        "0700-1700": (10, 0), "0730-1200/1700-2030": (8, 5),
        "700-1700": (10, 0), "730-1200/1700-2030": (8, 5), "1200-2130": (9.5, 0),
        "1000-2000": (10, 0), "1400-2030": (6.5, 0), "1400-2130": (7.5, 0),
    }

    jours_fr = {
        "Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi",
        "Thursday": "Jeudi", "Friday": "Vendredi", "Saturday": "Samedi", "Sunday": "Dimanche"
    }

    def normalize_horaire(horaire):
        if pd.isna(horaire): return ""
        horaire = str(horaire).strip().upper()
        horaire = horaire.replace(',', '.').replace(';', '.').replace(' ', '').replace('H', '')
        horaire = re.sub(r'[^0-9./\-]', '', horaire)
        return horaire.replace('.', '')

    def extract_hours(horaire):
        horaire = normalize_horaire(horaire)
        segments = re.split(r"/", horaire)
        start = segments[0].split("-")[0]
        end = segments[-1].split("-")[-1]
        return start, end

    def get_travail_coupure(horaire_normalise):
        horaire_normalise = horaire_normalise.replace(" ", "")
        for key in horaire_mapping:
            key_normalise = normalize_horaire(key).replace(" ", "")
            if horaire_normalise == key_normalise:
                return horaire_mapping[key]
        return 0, 0

    def horaire_to_hhmm(heure_str):
        if not heure_str or pd.isna(heure_str): return ""
        match = re.match(r"(\d{1,2})(\d{2})", heure_str)
        if match:
            h = int(match.group(1))
            m = int(match.group(2))
            return f"{h:02d}:{m:02d}"
        return ""

    def decimal_to_hhmm(decimal_val):
        try:
            total_minutes = round(float(decimal_val) * 60)
            h = total_minutes // 60
            m = total_minutes % 60
            return f"{h:02d}:{m:02d}"
        except:
            return "00:00"

    results = []
    for _, row in df_interim.iterrows():
        try:
            date_obj = pd.to_datetime(row["Date"], dayfirst=True)
            jour_fr = jours_fr[date_obj.strftime("%A")]
            date_str = f"{jour_fr} {date_obj.strftime('%d/%m/%Y')}"
        except:
            date_str = ""

        horaire = str(row["Horaire"])
        horaire_norm = normalize_horaire(horaire)
        heure_debut_raw, heure_fin_raw = extract_hours(horaire)
        t_travail, t_coupure = get_travail_coupure(horaire_norm)

        results.append({
            "Nom": "Interimaire",
            "Prénom": "Interimaire",
            "Poste": "Accompagnant éducatif et soc",
            "Stucture(s)": "Mas Montaines",
            "Date": date_str,
            "Heure de début de travail": horaire_to_hhmm(heure_debut_raw),
            "Temps de coupure": decimal_to_hhmm(t_coupure),
            "Heure de fin de travail": horaire_to_hhmm(heure_fin_raw),
            "Temps travaillé": decimal_to_hhmm(t_travail),
            "Personne remplacée": "",
            "Motif": "",
            "Info complémentaire sur le motif": "",
            "Unite(s)": f"Groupe {row['Groupe']}",
            "Précisez si coefficient EXTERNAT": "",
            "Commentaires": ""
        })

    df_badakan = pd.DataFrame(results)
    fichier_badakan = "badakan.csv"
    df_badakan.to_csv(fichier_badakan, sep=';', index=False, encoding='utf-8-sig')
    return fichier_badakan

# === Partie 3 : Génération du fichier Badakan ===
st.header("3️⃣ Génération du fichier Badakan")
uploaded_file_3 = st.file_uploader("Uploader le fichier avec l'onglet interimaire", type=["xlsx"], key="upload3")
if uploaded_file_3 and st.button("Générer le fichier Badakan"):
    interimaire_path = save_uploaded_file(uploaded_file_3, "interimaire")
    with st.spinner("Génération du fichier Badakan..."):
        fichier_badakan = traitement_partie3(interimaire_path)
    st.success("✅ Fichier Badakan généré !")
    st.download_button("📥 Télécharger Badakan.csv", data=open(fichier_badakan, "rb"), file_name="badakan.csv")