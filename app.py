import streamlit as st
import pandas as pd
import os
import tempfile
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy
import re
from datetime import datetime


# ✅ CECI DOIT ÊTRE LA PREMIÈRE COMMANDE STREAMLIT
st.set_page_config(page_title="Générateur de planning", layout="centered")
# === CONFIGURATION ===

PASSWORD = st.secrets["PLANNING_APP_PASSWORD"]



# === AUTHENTIFICATION ===
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        with st.form("Login"):
            pwd = st.text_input("🔐 Entrez le mot de passe", type="password")
            submitted = st.form_submit_button("Se connecter")
            if submitted and pwd == PASSWORD:
                st.session_state.authenticated = True
            elif submitted:
                st.error("Mot de passe incorrect")
    return st.session_state.authenticated

if not check_password():
    st.stop()

#st.set_page_config(page_title="Générateur de planning", layout="centered")
st.title("🗓️ Générateur de planning MAS Montaines")

# === Fonctions utilitaires génériques ===
def save_uploaded_file(uploaded_file, suffix):
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, f"{Path(uploaded_file.name).stem}_{suffix}{Path(uploaded_file.name).suffix}")
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return file_path
# ─────────────────────────────────────────────────────────────
# Helpers (à placer après les imports, avant les fonctions Streamlit)
# ─────────────────────────────────────────────────────────────
def nettoyer_nom_ligne4(ws, col_debut=4, col_fin=34):
    """
    Efface 'Nom/' (insensible à la casse/espaces) sur la ligne 4, 
    colonnes D→AH incluses.
    """
    for col in range(col_debut, col_fin + 1):      # 4 = D, 34 = AH
        cell = ws.cell(row=4, column=col)
        if isinstance(cell.value, str) and cell.value.strip().lower() == "nom/":
            cell.value = None
# ─────────────────────────────────────────────────────────────
# ------------------------------------------------------------------
# Mapping des 3 derniers chiffres vers le nom de la structure
# ------------------------------------------------------------------
import pandas as pd

def get_structure_mapping() -> dict[str, str]:
    data = [
        ("6750404", "EA ADAPAYSAGE BOURG"),
        ("6750405", "EA ADAPAYSAGE HAUT BUGEY"),
        ("6750309", "ESAT BELLEGARDE INDUSTRIE"),
        ("6750313", "ESAT CENTRE DE VIE RURALE"),
        ("6750307", "ESAT LA LECHERE"),
        ("6750305", "ESAT LE PENNESSUY"),
        ("6750311", "ESAT LES ATELIERS DE NIERME"),
        ("6750303", "ESAT LES BROSSES"),
        ("6750301", "ESAT LES DOMBES"),
        ("6750315", "ESAT LES TEPPES"),
        ("6750503", "FAM PRE LA TOUR"),
        ("6750504", "FAM SOUS LA ROCHE"),
        ("6750215", "FOYER BELLEVUE"),
        ("6750212", "FOYER DE TREFFORT"),
        ("6750213", "FOYER COURTES VERNOUX"),
        ("6750203", "FOYER CROIX BLANCHE"),
        ("6750201", "FOYER DE DOMAGNE"),
        ("6750210", "FOYER DE LASSIGNIEU"),
        ("6750207", "FOYER LE SOUS BOIS"),
        ("6750204", "FOYER LE VILLARDOIS"),
        ("6750202", "FOYER LES 4 VENTS"),
        ("6750209", "FOYER LES FLORALIES"),
        ("6750211", "FOYER LES PATIOS"),
        ("6750206", "FOYER LES PRES DE BROU"),
        ("6750214", "FOYER LES SOURDIERES"),
        ("6750208", "FOYER LE VAL FLEURI"),
        ("6750300", "CHAMP D'OR"),
        ("6750102", "IME GEORGES LOISEAU"),
        ("6750105", "IME L'ARMAILLOU"),
        ("6750101", "IME LE PRELION"),
        ("6750103", "IME LES SAPINS"),
        ("6750402", "EA DE BROU"),
        ("6750104", "IME SERVICE LES MUSCARIS"),
        ("6750401", "EA MAISONNETTE"),
        ("6750403", "EA MAISON DES PAYS DE L'AIN"),
        ("6750505", "MAS BELLEVUE"),
        ("6750502", "MAS LES MONTAINES"),
        ("6750501", "MAS MONTPLAISANT"),
        ("6750205", "SAVS LE PASSAGE BG EN B"),
        ("6750001", "ADAPEI DE L'AIN SIEGE SOCIAL"),
        ("6750007", "PCPE"),
        ("6750004", "POLE GEST BOURG EN BRESSE"),
        ("6750005", "POLE DE GESTION OYONNAX"),
        ("6750006", "POLE DE GESTION BELLEY"),
        ("6750003", "POLE GEST FONC TRANSVERSES"),
    ]
    df = pd.DataFrame(data, columns=["SIRH_ID", "SIRH_NOM"])
    df["Code"] = df["SIRH_ID"].str[-3:]
    return df.drop_duplicates("Code").set_index("Code")["SIRH_NOM"].to_dict()

STRUCTURE_MAP = get_structure_mapping()



def traitement_partie2(fichier_source):
    wb = load_workbook(fichier_source)
        # 1️⃣ : supprimer toutes les tables héritées du classeur source
    for ws in wb.worksheets:
        for tbl in list(ws._tables):
            ws.remove_table(tbl.name)

    # 2️⃣ : supprimer d'éventuels onglets "lecture" / "interimaire"
    for name in ["lecture", "interimaire"]:
        if name in wb.sheetnames:
            wb.remove(wb[name])
    #for ws in wb.worksheets:
    #    ws._tables = []
#
    #for name in ["lecture", "interimaire"]:
    #    if name in wb.sheetnames:
    #        std = wb[name]
    #        wb.remove(std)
#
    ws_source = wb.active
    colonnes = list(range(4, 35))
    mois_map = {
        "Jan": "01", "Fév": "02", "Fev": "02", "Mar": "03", "Mars": "03", "Avr": "04", "Mai": "05",
        "Juin": "06", "Jui": "07", "Juil": "07", "Août": "08", "Aou": "08", "Sep": "09", "Sept": "09",
        "Oct": "10", "Nov": "11", "Déc": "12", "Dec": "12"
    }

    dates_colonnes = {}
    for col in colonnes:
        val = ws_source.cell(row=1, column=col).value
        if isinstance(val, str) and "\n" in val:
            parts = val.strip().split("\n")
            if len(parts) == 2:
                jour, mois = parts
                jour = jour[1:] if jour and jour[0] in "LMJVSD" else jour
                if jour.isdigit():
                    dates_colonnes[col] = f"{int(jour):02d}/{mois_map.get(mois.strip().capitalize()[:4], '00')}/2025"

    lignes_donnees = []
    #for row in range(1, ws_source.max_row - 4):
    #    if ws_source.cell(row=row, column=3).value == "Hor.":
    #        ligne_hor = row
    #        ligne_lieu = row + 1
    #        ligne_act = row + 2
    #        ligne_nom = row + 3
    #        ligne_prenom = row + 4
#
    #        valeur_nom_colA = ws_source.cell(row=ligne_hor, column=1).value or ""
#
    #        for col in colonnes:
    #            val_act = ws_source.cell(row=ligne_act, column=col).value
    #            val_hor = ws_source.cell(row=ligne_hor, column=col).value
    #            val_lieu = ws_source.cell(row=ligne_lieu, column=col).value
#
    #            if isinstance(val_act, str) and val_act.startswith("502G"):
    #                groupe = val_act[-1] if val_act[-1].isdigit() else ""
    #                date = dates_colonnes.get(col, "")
    #                lignes_donnees.append([
    #                    date, groupe, val_hor, "", "", valeur_nom_colA.replace("\n", " "), val_lieu
    #                ])

       # ➡️ Au lieu de faire `for row in range(1, ws_source.max_row - 4):`
    # on va détecter **toutes** les lignes où col=3 vaut "Hor."
    max_row = ws_source.max_row
    lignes_hor = [
        r for r in range(1, max_row + 1)
        if ws_source.cell(row=r, column=3).value == "Hor."
    ]

    lignes_donnees = []
    for ligne_hor in lignes_hor:
        # on s'assure qu'on a bien la place pour lire les 4 lignes suivantes


        ligne_lieu   = ligne_hor + 1
        ligne_act    = ligne_hor + 2
        ligne_nom    = ligne_hor + 3
        ligne_prenom = ligne_hor + 4

            # ✔️ Si l’un des indices dépasse max_row, on l’ajuste au max
        if ligne_prenom > max_row:
            # On ne quitte plus la boucle : on prend ce qu’on peut
            ligne_lieu   = min(ligne_lieu,   max_row)
            ligne_act    = min(ligne_act,    max_row)
            ligne_nom    = min(ligne_nom,    max_row)
            ligne_prenom = min(ligne_prenom, max_row)

        valeur_nom_colA = ws_source.cell(row=ligne_hor, column=1).value or ""

        for col in colonnes:
            val_act = ws_source.cell(row=ligne_act, column=col).value
            val_hor = ws_source.cell(row=ligne_hor, column=col).value
            val_lieu = ws_source.cell(row=ligne_lieu, column=col).value

            #if isinstance(val_act, str) and val_act.startswith("502G"):
            #    groupe = val_act[-1] if val_act[-1].isdigit() else ""
            #    date   = dates_colonnes.get(col, "")
            #    lignes_donnees.append([
            #        date,
            #        groupe,
            #        val_hor,
            #        "",  # Motif
            #        "",  # NOM de la personne remplacée
            #        valeur_nom_colA.replace("\n", " "),
            #        val_lieu
            #    ])

            if isinstance(val_act, str) and re.match(r"\d{3}G\d", val_act):
                # ─── filtration horaire « 00:00 - 00:00 » ─────────────────────────
                if isinstance(val_hor, str):
                    hor_clean = val_hor.replace("\n", "").replace(" ", "")
                    if re.match(r"^0{1,2}:0{2}-0{1,2}:0{2}$", hor_clean):
                        continue  # on saute
    # -----------------------------------------------------------
    # 1) Code à 3 chiffres pour le titre          → '502'
    # 2) Groupe complet pour la colonne « Groupe » → '502G1'
    # -----------------------------------------------------------
                code_structure = val_act[:3]    # ex. '502'
                groupe_complet = val_act.strip()  # ex. '502G1'
                date_cell      = dates_colonnes.get(col, "")

                lignes_donnees.append([
                    date_cell,
                    groupe_complet,          # <─ colonne Groupe MAJ
                    val_hor,
                    "",                      # Motif
                    "",                      # NOM personne remplacée
                    valeur_nom_colA.replace("\n", " "),
                    val_lieu
                ])


    lignes_donnees = sorted(lignes_donnees, key=lambda x: (pd.to_datetime(x[0], dayfirst=True), x[1]))

    entetes = ["Date", "Groupe", "Horaire", "Motif", "NOM de la personne remplacée", "Nom", "Agence"]

    def creer_onglet(nom_onglet, lignes, code_structure="502"):
        """
        nom_onglet : 'lecture' ou 'interimaire'
        lignes     : données à insérer
        code_structure : '502', '404', etc.
        """
        nom_structure = STRUCTURE_MAP.get(code_structure, f"Structure {code_structure}")
        ws = wb.create_sheet(nom_onglet)

        # ▸ Titre fusionné ligne 3
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=9)
        cell_titre = ws.cell(row=3, column=3, value=nom_structure.upper())
        cell_titre.alignment = Alignment(horizontal="center")
        cell_titre.font = Font(name="Aptos Narrow", size=11, bold=True)
        cell_titre.fill = PatternFill("solid", start_color="FBE2D5")

        # ▸ En-têtes
        entetes = ["Date", "Groupe", "Horaire", "Motif",
                   "NOM de la personne remplacée", "Nom", "Agence"]
        for idx, val in enumerate(entetes, start=3):
            cell = ws.cell(row=4, column=idx, value=val)
            cell.font = Font(name="Aptos Narrow", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center")

        # ▸ Données
        ligne = 5
        for ligne_data in lignes:
            for col_index, val in enumerate(ligne_data, start=3):
                cell = ws.cell(row=ligne, column=col_index, value=val)
                cell.font = Font(name="Aptos Narrow", size=11)
                cell.alignment = Alignment(horizontal="center")
            ligne += 1

        # ▸ Table Excel (jolie)
        if lignes:
            ref = f"C4:I{ligne-1}"
            table = Table(displayName=f"Table{nom_onglet.capitalize()}", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

        # Largeur colonnes
        for col_idx in range(3, 10):
            ws.column_dimensions[chr(64+col_idx)].width = 28 if nom_onglet == "lecture" else 20

    # On récupère le premier code rencontré pour nommer l’onglet
    code_pour_titre = lignes_donnees[0][1][:3] if lignes_donnees else "502"

    #creer_onglet("lecture", lignes_donnees)
    creer_onglet("lecture", lignes_donnees, code_structure=code_pour_titre)


    df = pd.DataFrame(lignes_donnees, columns=entetes)
    df_interim = df[df["Agence"] == "A POURVOIR"].copy()
    df_interim["Date"] = pd.to_datetime(df_interim["Date"], dayfirst=True)
    df_interim = df_interim.sort_values(by=["Date", "Groupe"])
    df_interim["Date"] = df_interim["Date"].dt.strftime("%d/%m/%Y")
    creer_onglet("interimaire", df_interim.values.tolist())

    fichier_sortie = "planning_final_complet.xlsx"
    wb.save(fichier_sortie)
    return fichier_sortie

st.header("2️⃣ Génération des onglets 'lecture' et 'interimaire'")
uploaded_file_2 = st.file_uploader("Uploader le fichier planning stylisé modifié", type=["xlsx"], key="upload2")
if uploaded_file_2 and st.button("Créer les onglets Lecture & Interimaire"):
    stylized_path = save_uploaded_file(uploaded_file_2, "stylized")
    with st.spinner("Création des onglets Lecture et Interimaire..."):
        fichier_resultat = traitement_partie2(stylized_path)
    st.success("✅ Onglets créés !")
    st.download_button("📅 Télécharger le fichier modifié", data=open(fichier_resultat, "rb"), file_name=os.path.basename(fichier_resultat))


# === Partie 3 : Génération du fichier Badakan ===
def adapter_badakan(df_interim, structure_map):
    import re
    from datetime import datetime

    horaire_mapping = {
        "700-1430": (7.5, 0), "1400-2130": (7.5, 0), "800-2000": (12, 0), "0700-1430": (7.5, 0), "0800-2000": (12, 0),
        "0700-1700": (10, 0), "0730-1200/1700-2030": (8, 5),
        "700-1700": (10, 0), "730-1200/1700-2030": (8, 5), "1200-2130": (9.5, 0),
        "1000-2000": (10, 0), "1400-2030": (6.5, 0), "1400-2130": (7.5, 0),
    }

    jours_fr = {
        "Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi",
        "Thursday": "Jeudi", "Friday": "Vendredi", "Saturday": "Samedi", "Sunday": "Dimanche"
    }

    def normalize_horaire(horaire):
        if pd.isna(horaire): return ""
        horaire = str(horaire).strip().upper()
        horaire = horaire.replace(',', '.').replace(';', '.').replace(' ', '').replace('H', '')
        horaire = re.sub(r'[^0-9./\-]', '', horaire)
        return horaire.replace('.', '')

    def extract_hours(horaire):
        horaire = normalize_horaire(horaire)
        segments = re.split(r"/", horaire)
        start = segments[0].split("-")[0]
        end = segments[-1].split("-")[-1]
        return start, end

    def get_travail_coupure(horaire_normalise):
        horaire_normalise = horaire_normalise.replace(" ", "")
        for key in horaire_mapping:
            key_normalise = normalize_horaire(key).replace(" ", "")
            if horaire_normalise == key_normalise:
                return horaire_mapping[key]
        return 0, 0

    def horaire_to_hhmm(heure_str):
        if not heure_str or pd.isna(heure_str): return ""
        match = re.match(r"(\d{1,2})(\d{2})", heure_str)
        if match:
            h = int(match.group(1))
            m = int(match.group(2))
            return f"{h:02d}:{m:02d}"
        return ""

    def decimal_to_hhmm(decimal_val):
        try:
            total_minutes = round(float(decimal_val) * 60)
            h = total_minutes // 60
            m = total_minutes % 60
            return f"{h:02d}:{m:02d}"
        except:
            return "00:00"

    results = []
    for _, row in df_interim.iterrows():
        groupe_val = str(row["Groupe"])
        code_struct = groupe_val[:3] if len(groupe_val) >= 4 else "000"
        unite_val = groupe_val[3:] if len(groupe_val) > 3 else ""
        structure = structure_map.get(code_struct, f"Structure {code_struct}")

        try:
            date_obj = pd.to_datetime(row["Date"], dayfirst=True)
            jour_fr = jours_fr[date_obj.strftime("%A")]
            date_str = f"{jour_fr} {date_obj.strftime('%d/%m/%Y')}"
        except:
            date_str = ""

        horaire = str(row["Horaire"])
        horaire_norm = normalize_horaire(horaire)
        heure_debut_raw, heure_fin_raw = extract_hours(horaire)
        t_travail, t_coupure = get_travail_coupure(horaire_norm)

        results.append({
            "Nom": "Interimaire",
            "Prénom": "Interimaire",
            "Poste": "Accompagnant éducatif et soc",
            "Stucture(s)": structure,
            "Date": date_str,
            "Heure de début de travail": horaire_to_hhmm(heure_debut_raw),
            "Temps de coupure": decimal_to_hhmm(t_coupure),
            "Heure de fin de travail": horaire_to_hhmm(heure_fin_raw),
            "Temps travaillé": decimal_to_hhmm(t_travail),
            "Personne remplacée": row.get("Personne remplacée", ""),
            "Motif": row.get("Motif", ""),
            "Info complémentaire sur le motif": "",
            "Unite(s)": unite_val,
            "Précisez si coefficient EXTERNAT": "",
            "Commentaires": ""
        })

    return pd.DataFrame(results)

#def traitement_partie3(fichier_interimaire):
    df_interim = pd.read_excel(fichier_interimaire, sheet_name="interimaire", skiprows=3, usecols="C:I")
    df_interim.columns = ["Date", "Groupe", "Horaire", "Motif", "Personne remplacée", "Nom", "Agence"]


    horaire_mapping = {
        "700-1430": (7.5, 0), "1400-2130": (7.5, 0), "800-2000": (12, 0), "0700-1430": (7.5, 0),  "0800-2000": (12, 0),
        "0700-1700": (10, 0), "0730-1200/1700-2030": (8, 5),
        "700-1700": (10, 0), "730-1200/1700-2030": (8, 5), "1200-2130": (9.5, 0),
        "1000-2000": (10, 0), "1400-2030": (6.5, 0), "1400-2130": (7.5, 0),
    }

    jours_fr = {
        "Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi",
        "Thursday": "Jeudi", "Friday": "Vendredi", "Saturday": "Samedi", "Sunday": "Dimanche"
    }

    def normalize_horaire(horaire):
        if pd.isna(horaire): return ""
        horaire = str(horaire).strip().upper()
        horaire = horaire.replace(',', '.').replace(';', '.').replace(' ', '').replace('H', '')
        horaire = re.sub(r'[^0-9./\-]', '', horaire)
        return horaire.replace('.', '')

    def extract_hours(horaire):
        horaire = normalize_horaire(horaire)
        segments = re.split(r"/", horaire)
        start = segments[0].split("-")[0]
        end = segments[-1].split("-")[-1]
        return start, end

    def get_travail_coupure(horaire_normalise):
        horaire_normalise = horaire_normalise.replace(" ", "")
        for key in horaire_mapping:
            key_normalise = normalize_horaire(key).replace(" ", "")
            if horaire_normalise == key_normalise:
                return horaire_mapping[key]
        return 0, 0

    def horaire_to_hhmm(heure_str):
        if not heure_str or pd.isna(heure_str): return ""
        match = re.match(r"(\d{1,2})(\d{2})", heure_str)
        if match:
            h = int(match.group(1))
            m = int(match.group(2))
            return f"{h:02d}:{m:02d}"
        return ""

    def decimal_to_hhmm(decimal_val):
        try:
            total_minutes = round(float(decimal_val) * 60)
            h = total_minutes // 60
            m = total_minutes % 60
            return f"{h:02d}:{m:02d}"
        except:
            return "00:00"

    results = []
    for _, row in df_interim.iterrows():
        try:
            date_obj = pd.to_datetime(row["Date"], dayfirst=True)
            jour_fr = jours_fr[date_obj.strftime("%A")]
            date_str = f"{jour_fr} {date_obj.strftime('%d/%m/%Y')}"
        except:
            date_str = ""

        horaire = str(row["Horaire"])
        horaire_norm = normalize_horaire(horaire)
        heure_debut_raw, heure_fin_raw = extract_hours(horaire)
        t_travail, t_coupure = get_travail_coupure(horaire_norm)

        results.append({
            "Nom": "Interimaire",
            "Prénom": "Interimaire",
            "Poste": "Accompagnant éducatif et soc",
            "Stucture(s)": "Mas Montaines",
            "Date": date_str,
            "Heure de début de travail": horaire_to_hhmm(heure_debut_raw),
            "Temps de coupure": decimal_to_hhmm(t_coupure),
            "Heure de fin de travail": horaire_to_hhmm(heure_fin_raw),
            "Temps travaillé": decimal_to_hhmm(t_travail),
            "Personne remplacée": "",
            "Motif": "",
            "Info complémentaire sur le motif": "",
            "Unite(s)": f"Groupe {row['Groupe']}",
            "Précisez si coefficient EXTERNAT": "",
            "Commentaires": ""
        })

    #df_badakan = pd.DataFrame(results)
    #fichier_badakan = "badakan.csv"
    #df_badakan.to_csv(fichier_badakan, sep=';', index=False, encoding='utf-8-sig')
    #return fichier_badakan
    # Nouvelle génération propre avec mapping et champs dynamiques
    #df_badakan = adapter_badakan(df_interim, STRUCTURE_MAP)
    df_badakan = adapter_badakan_version_auto(df_interim, STRUCTURE_MAP)

    fichier_badakan = "badakan.csv"
    df_badakan.to_csv(fichier_badakan, sep=';', index=False, encoding='utf-8-sig')
    return fichier_badakan
def adapter_badakan_version_auto(df_interim, structure_map):
    import re
    from datetime import datetime

    jours_fr = {
        "Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi",
        "Thursday": "Jeudi", "Friday": "Vendredi", "Saturday": "Samedi", "Sunday": "Dimanche"
    }

    def calculer_infos_horaires(horaire_str):
        from datetime import datetime
        def parse_time(h): return datetime.strptime(h.strip(), "%H:%M")

        if not isinstance(horaire_str, str):
            return "", "", "00:00", 0.0

        horaire_str = horaire_str.replace("\n", "").strip()
        segments = [s.strip() for s in horaire_str.split("/") if s.strip()]
        plages = [tuple(seg.split("-")) for seg in segments if "-" in seg]

        if not plages:
            return "", "", "00:00", 0.0

        heure_debut = plages[0][0].strip()
        heure_fin   = plages[-1][1].strip()

        debut_dt = parse_time(heure_debut)
        fin_dt   = parse_time(heure_fin)

        total_travail = sum((parse_time(f) - parse_time(d)).seconds for d, f in plages) // 60  # minutes
        total_cadre   = (fin_dt - debut_dt).seconds // 60
        coupure_min   = total_cadre - total_travail
        temps_decimal = round(total_travail / 60, 2)

        def min_to_hhmm(m):
            h, mn = divmod(m, 60)
            return f"{h:02d}:{mn:02d}"

        return heure_debut, heure_fin, min_to_hhmm(coupure_min), temps_decimal


    results = []
    for _, row in df_interim.iterrows():
        groupe_val = str(row["Groupe"])
        code_struct = groupe_val[:3] if len(groupe_val) >= 4 else "000"
        unite_val = groupe_val[3:] if len(groupe_val) > 3 else ""
        structure = structure_map.get(code_struct, f"Structure {code_struct}")

        try:
            date_obj = pd.to_datetime(row["Date"], dayfirst=True)
            jour_fr = jours_fr[date_obj.strftime("%A")]
            date_str = f"{jour_fr} {date_obj.strftime('%d/%m/%Y')}"
        except:
            date_str = ""

        horaire = str(row["Horaire"])
        heure_debut, heure_fin, coupure, t_travail = calculer_infos_horaires(horaire)

        results.append({
            "Nom": "Interimaire",
            "Prénom": "Interimaire",
            "Poste": "Accompagnant éducatif et soc",
            "Stucture(s)": structure,
            "Date": date_str,
            "Heure de début de travail": heure_debut,
            "Temps de coupure": coupure,
            "Heure de fin de travail": heure_fin,
            "Temps travaillé": str(t_travail).replace(".", ","),
            "Personne remplacée": row.get("Personne remplacée", ""),
            "Motif": row.get("Motif", ""),
            "Info complémentaire sur le motif": "",
            "Unite(s)": unite_val,
            "Précisez si coefficient EXTERNAT": "",
            "Commentaires": ""
        })

    return pd.DataFrame(results)

def traitement_partie3(fichier_interimaire):
    df_interim = pd.read_excel(
        fichier_interimaire,
        sheet_name="interimaire",
        skiprows=3,
        usecols="C:I"
    )
    df_interim.columns = [
        "Date", "Groupe", "Horaire", "Motif",
        "Personne remplacée", "Nom", "Agence"
    ]

    # Utilise la version avec analyse automatique des horaires
    df_badakan = adapter_badakan_version_auto(df_interim, STRUCTURE_MAP)

    fichier_badakan = "badakan.csv"
    df_badakan.to_csv(fichier_badakan, sep=';', index=False, encoding='utf-8-sig')
    return fichier_badakan

# === Partie 3 : Génération du fichier Badakan ===
st.header("3️⃣ Génération du fichier Badakan")
uploaded_file_3 = st.file_uploader("Uploader le fichier avec l'onglet interimaire", type=["xlsx"], key="upload3")
if uploaded_file_3 and st.button("Générer le fichier Badakan"):
    interimaire_path = save_uploaded_file(uploaded_file_3, "interimaire")
    with st.spinner("Génération du fichier Badakan..."):
        fichier_badakan = traitement_partie3(interimaire_path)
    st.success("✅ Fichier Badakan généré !")
    st.download_button("📥 Télécharger Badakan.csv", data=open(fichier_badakan, "rb"), file_name="badakan.csv")