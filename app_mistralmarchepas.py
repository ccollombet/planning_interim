import streamlit as st
import pandas as pd
from datetime import datetime
import re
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

# === CONFIGURATION ===
PASSWORD = "Celine01$"

horaire_mapping = {
    "700-1430": (7.5, 0),
    "1400-2130": (7.5, 0),
    "800-2000": (12, 0),
    "700-1700": (10, 0),
    "730-1200/1700-2030": (8, 5),
    "1200-2130": (9.5, 0),
    "1000-2000": (10, 0),
    "1400-2030": (6.5, 0),
    "1400-2130": (7.5, 0),
}

jours_fr = {
    "Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi",
    "Thursday": "Jeudi", "Friday": "Vendredi",
    "Saturday": "Samedi", "Sunday": "Dimanche"
}

# === UTILITAIRES ===
def normalize_horaire(horaire):
    if pd.isna(horaire): return ""
    horaire = str(horaire).strip().upper()
    horaire = horaire.replace(',', '.').replace(';', '.').replace(' ', '').replace('H', '')
    horaire = re.sub(r'[^0-9./\-]', '', horaire)
    return horaire.replace('.', '')

def extract_hours(horaire):
    horaire = normalize_horaire(horaire)
    segments = re.split(r"/", horaire)
    start = segments[0].split("-")[0]
    end = segments[-1].split("-")[-1]
    return start, end

def get_travail_coupure(horaire_normalise):
    horaire_normalise = horaire_normalise.replace(" ", "")
    for key in horaire_mapping:
        key_normalise = normalize_horaire(key).replace(" ", "")
        if horaire_normalise == key_normalise:
            return horaire_mapping[key]
    return 0, 0

def horaire_to_hhmm(heure_str):
    if not heure_str or pd.isna(heure_str): return ""
    match = re.match(r"(\d{1,2})(\d{2})", heure_str)
    if match:
        h = int(match.group(1))
        m = int(match.group(2))
        return f"{h:02d}:{m:02d}"
    return ""

def decimal_to_hhmm(decimal_val):
    try:
        total_minutes = round(float(decimal_val) * 60)
        h = total_minutes // 60
        m = total_minutes % 60
        return f"{h:02d}:{m:02d}"
    except:
        return "00:00"

# === AUTHENTIFICATION ===
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        with st.form("Login"):
            pwd = st.text_input("🔐 Entrez le mot de passe", type="password")
            submitted = st.form_submit_button("Se connecter")
            if submitted and pwd == PASSWORD:
                st.session_state.authenticated = True
            elif submitted:
                st.error("Mot de passe incorrect")

    return st.session_state.authenticated

# === LOGIQUE DE TRAITEMENT ===
def traiter_fichier(df, structure_name, groupe_prefix):
    results = []
    groupe_courant = None

    for row in range(1, df.shape[0]):
        valeur_colA = df.iloc[row, 0]
        valeur_colB = str(df.iloc[row, 1]).strip().upper()

        if pd.notna(valeur_colA) and valeur_colA != '':
            try:
                if float(valeur_colA).is_integer():
                    groupe_courant = str(int(float(valeur_colA)))
                else:
                    groupe_courant = str(valeur_colA)
            except ValueError:
                groupe_courant = str(valeur_colA).strip()

        if valeur_colB == "NOM" and groupe_courant:
            for col in range(2, df.shape[1]):
                nom = str(df.iloc[row, col]).strip()
                agence = str(df.iloc[row + 1, col]).strip()
                horaire = str(df.iloc[row - 1, col]).strip()

                try:
                    day = int(df.iloc[0, col])
                    date_obj = datetime(2025, 6, day)
                    jour_fr = jours_fr[date_obj.strftime("%A")]
                    date_str = f"{jour_fr} {date_obj.strftime('%d/%m/%Y')}"
                except:
                    continue

                if nom.upper() == "INTERIMAIRE":
                    heure_debut_raw, heure_fin_raw = extract_hours(horaire)
                    horaire_normalise = normalize_horaire(horaire)
                    t_travail, t_coupure = get_travail_coupure(horaire_normalise)

                    results.append({
                        "Nom": "Interimaire",
                        "Prénom": "Interimaire",
                        "Poste": "Accompagnant éducatif et soc",
                        "Stucture(s)": structure_name,
                        "Date": date_str,
                        "Heure de début de travail": horaire_to_hhmm(heure_debut_raw),
                        "Temps de coupure": decimal_to_hhmm(t_coupure),
                        "Heure de fin de travail": horaire_to_hhmm(heure_fin_raw),
                        "Temps travaillé": decimal_to_hhmm(t_travail),
                        "Personne remplacée": "",
                        "Motif": "",
                        "Info complémentaire sur le motif": f"Groupe {groupe_courant}",
                        "Unite(s)": "",
                        "Précisez si coefficient EXTERNAT": "",
                        "Commentaires": ""
                    })
    return pd.DataFrame(results)

def generate_red_file(uploaded_file):
    df = pd.read_excel(uploaded_file, sheet_name=0, header=None)

    # === DÉTECTION DU MOIS ===
    mois_texte = None
    for cell in df.iloc[0, 3:]:
        if isinstance(cell, str) and "\n" in cell:
            parts = cell.split("\n")
            if len(parts) == 2:
                mois_texte = parts[1].strip()
                break

    mois_numerique = {
        "janv": 1, "févr": 2, "mars": 3, "avr": 4,
        "mai": 5, "juin": 6, "juil": 7, "août": 8,
        "sept": 9, "oct": 10, "nov": 11, "déc": 12
    }
    mois_clean = mois_texte.lower().replace(".", "")[:4] if mois_texte else "avr"
    mois = mois_numerique.get(mois_clean, 4)
    annee = 2025

    # === NETTOYAGE ===
    df = df[~df.apply(lambda row: str(row[0])[:10].count("/") >= 2 and ":" in str(row[0]), axis=1)].reset_index(drop=True)

    # === FORMAT DES COLONNES ===
    raw_headers = df.iloc[0].fillna("").astype(str)
    df.columns = [h.replace("\n", " ").strip() for h in raw_headers]
    df = df.drop(index=0).reset_index(drop=True)

    jours = []
    for cell in df.columns[3:]:
        jour_num = ''.join(filter(str.isdigit, str(cell)))
        if jour_num.isdigit():
            date_complete = pd.Timestamp(day=int(jour_num), month=mois, year=annee)
            jours.append(date_complete.strftime("%d/%m"))

    # === EXTRACTION CSV INTERMÉDIAIRE ===
    debug_rows = []
    for i in range(0, len(df), 8):
        bloc = df.iloc[i:i+8]
        if bloc.empty or pd.isna(bloc.iloc[0, 0]):
            continue
        nom_complet = str(bloc.iloc[0, 0]).replace("\n", " ").strip()
        maj_only = re.findall(r'\b[A-ZÉÈÇÂÊÎÔÛÄËÏÖÜÀÙ]+\b', nom_complet)
        nom = " ".join(maj_only)
        lieu_row = bloc.iloc[2, 3:]
        act_jour_row = bloc.iloc[3, 3:]
        try:
            u0040 = bloc[bloc.iloc[:, 2] == "U0040"].iloc[0, 3:].tolist()
            u0042 = bloc[bloc.iloc[:, 2] == "U0042"].iloc[0, 3:].tolist()
            u0044 = bloc[bloc.iloc[:, 2] == "U0044"].iloc[0, 3:].tolist()
            u0046 = bloc[bloc.iloc[:, 2] == "U0046"].iloc[0, 3:].tolist()
        except IndexError:
            continue
        for j, jour in enumerate(jours):
            if j >= len(u0040) or j >= len(u0046):
                continue
            debug_rows.append({
                "Nom": nom,
                "Jour": jour,
                "Act. jour": act_jour_row.iloc[j] if j < len(act_jour_row) else "",
                "U0040": u0040[j],
                "U0042": u0042[j],
                "U0044": u0044[j],
                "U0046": u0046[j],
                "Lieu": lieu_row.iloc[j] if j < len(lieu_row) else ""
            })

    # Suppression des doublons
    df_debug = pd.DataFrame(debug_rows)
    df_debug = df_debug.drop_duplicates(subset=['Nom', 'Jour'])

    # === CONSTRUCTION DU FICHIER FINAL ===
    def format_heure(valeur):
        if isinstance(valeur, str):
            valeur = valeur.replace(",", ".")
        try:
            valeur_float = float(valeur)
            heures = int(valeur_float)
            minutes = round((valeur_float - heures) * 100)
            return f"{heures:02d}h{minutes:02d}"
        except:
            return ""

    df_flat = df_debug.copy()
    df_flat = df_flat[(df_flat["Act. jour"].notna()) & (df_flat["U0040"].notna()) & (df_flat["U0046"].notna())]
    df_flat["Groupe"] = df_flat["Act. jour"].astype(str).str.upper().str.replace(" ", "")
    df_flat["Horaire"] = df_flat.apply(lambda row: f"{format_heure(row['U0040'])}–{format_heure(row['U0046'])}", axis=1)
    df_flat = df_flat[["Jour", "Groupe", "Nom", "Horaire", "Lieu"]]
    df_flat = df_flat.rename(columns={"Lieu": "Agence"})

    groupe_noms = {
        "501A": "AZUR", "501B": "BORA", "501C": "CARAÏBES",
        "501D": "DJERBA", "501E": "EGEE", "501F": "FIDJI"
    }
    df_flat["Groupe"] = df_flat["Groupe"].apply(lambda x: groupe_noms.get(x, x))

    # Filtrer les groupes
    groupes_a_afficher = ["AZUR", "BORA", "CARAÏBES", "DJERBA", "EGEE", "FIDJI"]
    df_flat = df_flat[df_flat["Groupe"].isin(groupes_a_afficher)]

    def heure_debut(horaire):
        match = re.search(r"(\d{2})h(\d{2})", str(horaire))
        if match:
            return int(match.group(1)) * 60 + int(match.group(2))
        return 9999

    df_flat["HeureTri"] = df_flat["Horaire"].apply(heure_debut)
    jours_uniques = sorted(df_flat["Jour"].unique(), key=lambda x: int(x[:2]))
    groupes_uniques = sorted(df_flat["Groupe"].unique())

    output_data = {"A": [], "B": []}
    for jour in jours_uniques:
        output_data[jour] = []

    for groupe in groupes_uniques:
        output_data["A"].extend([groupe] + [""] * 17)
        output_data["B"].extend(["Horaire", "NOM", "Agence"] * 6)
        df_grp = df_flat[df_flat["Groupe"] == groupe]
        for jour in jours_uniques:
            df_jour = df_grp[df_grp["Jour"] == jour].sort_values(by="HeureTri").reset_index(drop=True)
            horaires = ["" for _ in range(6)]
            noms = ["" for _ in range(6)]
            agences = ["" for _ in range(6)]
            for i in range(min(6, len(df_jour))):
                horaires[i] = df_jour.loc[i, "Horaire"]
                noms[i] = df_jour.loc[i, "Nom"]
                agences[i] = df_jour.loc[i, "Agence"]
            for i in range(6):
                output_data[jour].append(horaires[i])
                output_data[jour].append(noms[i])
                output_data[jour].append(agences[i])

    df_final = pd.DataFrame(output_data)

    # Charger le fichier Excel généré
    wb = load_workbook(io.BytesIO(df_final.to_excel(index=False)))
    ws = wb.active

    # Définir les styles
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Supprimer les en-têtes des colonnes A et B
    ws['A1'] = ""
    ws['B1'] = ""

    # Ajouter une ligne verte
    ws.insert_rows(2)
    for cell in ws[2]:
        cell.fill = green_fill

    # Fusionner les cellules pour les noms de groupe
    group_ranges = {
        "AZUR": (3, 19),
        "BORA": (20, 36),
        "CARAÏBES": (37, 53),
        "DJERBA": (54, 70),
        "EGEE": (71, 87),
        "FIDJI": (88, 104)
    }

    for group, (start_row, end_row) in group_ranges.items():
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1, value=group).alignment = Alignment(horizontal="center", vertical="center")

    # Appliquer des bordures fines
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Sauvegarder le fichier modifié
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()

    return excel_bytes

# === INTERFACE ===
if not check_password():
    st.stop()

st.title("Application de Gestion de Planning")

# Section 1: Génération du fichier ROUGE
st.header("1. Générer le fichier ROUGE")
uploaded_file_red = st.file_uploader("Déposez un fichier Excel pour le fichier ROUGE", type="xlsx", key="red_file_uploader")

if uploaded_file_red is not None:
    excel_bytes = generate_red_file(uploaded_file_red)
    st.download_button(
        label="📥 Télécharger le fichier ROUGE",
        data=excel_bytes,
        file_name="planning_rouge.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Section 2: Conversion pour Badakan
st.header("2. 🧾 Convertisseur de planning en CSV avec choix de la MAS")
choix_mas = st.radio("Choisissez la MAS :", ["Montaines", "Montplaisant"])
uploaded_file = st.file_uploader("Déposez un fichier Excel du planning", type="xlsx", key="csv_file_uploader")

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file, sheet_name=0, header=None)

    if choix_mas == "Montaines":
        df_result = traiter_fichier(df, "Mas Montaines", groupe_prefix="")
    else:
        df_result = traiter_fichier(df, "Mas Montplaisant", groupe_prefix="")

    if not df_result.empty:
        csv_buffer = io.StringIO()
        df_result.to_csv(csv_buffer, sep=';', index=False, encoding='utf-8-sig')
        csv_bytes = csv_buffer.getvalue().encode('utf-8-sig')

        st.success("✅ Conversion terminée avec succès !")
        st.download_button(
            label="📥 Télécharger le fichier CSV",
            data=csv_bytes,
            file_name="planning_interimaires_converti.csv",
            mime="text/csv"
        )
    else:
        st.warning("Le fichier n'a pas produit de lignes exploitables.")
